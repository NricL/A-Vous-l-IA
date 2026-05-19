"""
Ingestion de documents : chargement, découpage, ajout au vector store (Chroma via Haystack).
Pour les XLSX : indexation par ligne (1 ligne = 1 chunk), contenu = colonne rag_text_auto, reste en métadonnées.
"""

from pathlib import Path

from langchain_community.document_loaders import PyPDFLoader, TextLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_core.documents import Document as LangChainDocument
from openpyxl import load_workbook
from haystack import Document as HaystackDocument

from app.config import get_settings
from app.haystack_rag import index_documents_haystack

# Colonne dont la valeur est vectorisée ; les autres colonnes vont en métadonnées.
RAG_TEXT_AUTO_COLUMN = "rag_text_auto"


def _cell_str(value) -> str:
    """Convertit une cellule en chaîne pour contenu ou métadonnées."""
    if value is None:
        return ""
    return str(value).strip()


def _load_xlsx(file_path: str) -> list[LangChainDocument]:
    """
    Charge un XLSX : une seule feuille, première ligne = en-tête.
    Une ligne de données = un document : contenu = colonne rag_text_auto, autres colonnes en meta.
    Lignes vides (rag_text_auto vide) ignorées.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    sheet_name = sheet.title
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [_cell_str(c) for c in rows[0]]
    try:
        content_col_idx = headers.index(RAG_TEXT_AUTO_COLUMN)
    except ValueError:
        raise ValueError(
            f"Colonne '{RAG_TEXT_AUTO_COLUMN}' introuvable dans la première ligne. En-têtes : {headers}"
        )

    docs = []
    for row_idx, row in enumerate(rows[1:], start=2):  # ligne 2 = première ligne de données
        content = _cell_str(row[content_col_idx] if content_col_idx < len(row) else None)
        if not content:
            continue

        meta = {"source_file": Path(file_path).name, "sheet": sheet_name, "row_index": row_idx}
        for col_idx, header in enumerate(headers):
            if not header or col_idx == content_col_idx:
                continue
            value = row[col_idx] if col_idx < len(row) else None
            meta[header] = _cell_str(value)

        docs.append(LangChainDocument(page_content=content, metadata=meta))
    return docs


def get_loader_for_path(file_path: str):
    """Retourne le loader LangChain adapté au type de fichier (ou None pour xlsx, géré à part)."""
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return PyPDFLoader(file_path)
    if suffix in (".txt", ".md"):
        return TextLoader(file_path, encoding="utf-8")
    if suffix == ".xlsx":
        return None  # géré par load_and_split_documents
    raise ValueError(f"Type de fichier non supporté : {suffix}")


def load_and_split_documents(file_path: str) -> list[LangChainDocument]:
    """Charge un fichier et le découpe en chunks (documents LangChain). XLSX : 1 ligne = 1 doc, pas de découpage."""
    settings = get_settings()
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _load_xlsx(file_path)
    loader = get_loader_for_path(file_path)
    docs = loader.load()
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.chunk_size,
        chunk_overlap=settings.chunk_overlap,
        length_function=len,
    )
    return splitter.split_documents(docs)


def _lc_to_haystack_docs(lc_docs: list[LangChainDocument], file_path: str, metadata: dict | None = None) -> list[HaystackDocument]:
    """Convertit des documents LangChain en documents Haystack."""
    out = []
    for d in lc_docs:
        meta = dict(d.metadata or {})
        meta["source_file"] = str(Path(file_path).name)
        if metadata:
            meta.update(metadata)
        out.append(HaystackDocument(content=d.page_content, meta=meta))
    return out


def ingest_file(file_path: str, metadata: dict | None = None) -> list[str]:
    """
    Ingère un fichier dans Chroma via Haystack (embedding + écriture).
    Retourne une liste d'ids factices pour compatibilité API (count = len(ids)).
    """
    docs = load_and_split_documents(file_path)
    if not docs:
        return []
    for d in docs:
        d.metadata = d.metadata or {}
        d.metadata["source_file"] = str(Path(file_path).name)
        if metadata:
            d.metadata.update(metadata)
    haystack_docs = _lc_to_haystack_docs(docs, file_path, metadata)
    count = index_documents_haystack(haystack_docs)
    return [str(i) for i in range(count)]


def ingest_bytes(content: bytes, filename: str, metadata: dict | None = None) -> list[str]:
    """
    Ingère du contenu binaire (upload) dans Chroma.
    Écrit temporairement sur disque pour les loaders qui lisent des fichiers.
    """
    import tempfile
    suffix = Path(filename).suffix.lower()
    if suffix not in (".pdf", ".txt", ".md", ".xlsx"):
        raise ValueError(f"Type non supporté : {suffix}. Utilisez .pdf, .txt, .md ou .xlsx")
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
        f.write(content)
        path = f.name
    try:
        return ingest_file(path, metadata=metadata or {"filename": filename})
    finally:
        Path(path).unlink(missing_ok=True)
