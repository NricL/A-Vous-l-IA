#!/usr/bin/env python3
"""
Diagnostic Chroma pour un id de cas (ex. UC-0764) sans importer haystack_rag (évite chromadb).

Lit embedding_metadata via sqlite3, reconstruit un dict meta plat, applique la même logique
que _case_extra_fields_from_meta (alias CASE_EXTRA_FIELD_ALIASES).

Usage (depuis backend/) :
  PYTHONPATH=. python3 scripts/chroma_inspect_uc0764.py
  PYTHONPATH=. python3 scripts/chroma_inspect_uc0764.py --needle UC-0764
  CHROMA_PERSIST_DIR=/chemin/vers/chroma PYTHONPATH=. python3 scripts/chroma_inspect_uc0764.py
"""
from __future__ import annotations

import argparse
import logging
import os
import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger("chroma_uc")

from app.rag_constants import CASE_EXTRA_FIELD_ALIASES, STRUCTURED_VERBATIM_USER_KEYS  # noqa: E402


def _meta_first_raw_value(meta: dict, keys: tuple[str, ...]) -> str | None:
    for key in keys:
        if key not in meta:
            continue
        raw = meta[key]
        if raw is None:
            continue
        return raw if isinstance(raw, str) else str(raw)
    return None


def _case_extra_fields_from_meta(meta: dict | None) -> dict[str, str | None]:
    meta = meta or {}
    out: dict[str, str | None] = {}
    for canonical, aliases in CASE_EXTRA_FIELD_ALIASES.items():
        out[canonical] = _meta_first_raw_value(meta, aliases)
    return out


def _load_flat_meta_for_embedding(con: sqlite3.Connection, embedding_row_id: int) -> dict[str, str]:
    rows = con.execute(
        "SELECT key, string_value FROM embedding_metadata WHERE id=? AND string_value IS NOT NULL",
        (embedding_row_id,),
    ).fetchall()
    flat: dict[str, str] = {}
    for k, v in rows:
        nk = k[5:] if k.startswith("meta.") else k
        flat[nk] = v
    return flat


def _find_embedding_ids(con: sqlite3.Connection, needle: str) -> list[int]:
    return [
        r[0]
        for r in con.execute(
            "SELECT DISTINCT id FROM embedding_metadata WHERE string_value LIKE '%' || ? || '%'",
            (needle,),
        ).fetchall()
    ]


def _print_xlsx_headers(path: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.warning("openpyxl absent : impossible de lire %s", path)
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    wb.close()
    headers = [str(c).strip() if c is not None else "" for c in row]
    log.info("Feuille %r — %s en-têtes (premiers 40) :", sheet.title, len(headers))
    for i, h in enumerate(headers[:40], start=1):
        log.info("  %s. %r", i, h)


def _chroma_sqlite_path() -> Path:
    env_dir = os.environ.get("CHROMA_PERSIST_DIR")
    if env_dir:
        return Path(env_dir).resolve() / "chroma.sqlite3"
    here = Path(__file__).resolve().parents[1]
    return (here / "data" / "chroma" / "chroma.sqlite3").resolve()


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--needle", default="UC-0764", help="Sous-chaîne à chercher dans les meta (ex. UC-0764)")
    p.add_argument("--xlsx", type=Path, help="Si vide après alias : afficher les en-têtes Sheet1 de ce fichier")
    args = p.parse_args()
    needle = args.needle

    db_path = _chroma_sqlite_path()
    if not db_path.is_file():
        log.error("SQLite Chroma introuvable : %s", db_path)
        return 1

    log.info("SQLite : %s", db_path)
    con = sqlite3.connect(str(db_path))
    eids = _find_embedding_ids(con, needle)
    if not eids:
        log.error("Aucun embedding_metadata ne contient %r", needle)
        con.close()
        return 2
    eid = eids[0]
    log.info("Embedding row id(s) : %s (on utilise le premier)", eids[:5])

    flat = _load_flat_meta_for_embedding(con, eid)
    con.close()

    resolved = _case_extra_fields_from_meta(flat)
    four = STRUCTURED_VERBATIM_USER_KEYS[:4]
    log.info("")
    log.info("=== 4 premières clés (STRUCTURED_VERBATIM_USER_KEYS) — résolution alias ===")
    for k in four:
        v = resolved.get(k)
        log.info("%s : len=%s", k, len(v) if v else 0)
        log.info("  %r", (v[:500] + "…") if v and len(v) > 500 else v)

    pa = resolved.get("premiere_action_48h")
    log.info("")
    log.info("=== premiere_action_48h (non vide = alias OK ; comparer longueur au source) ===")
    log.info("longueur : %s", len(pa) if pa else 0)
    if pa:
        log.info("valeur complète :\n%s", pa)
    else:
        log.warning("VIDE : vérifier l’en-tête XLSX réel et le placer en tête du tuple premiere_action_48h dans rag_constants.py")
        log.info("Alias actuels : %s", CASE_EXTRA_FIELD_ALIASES.get("premiere_action_48h"))
        log.info("Clés meta candidates (48h / premiere / action / étape) :")
        for k in sorted(flat.keys()):
            lk = k.lower()
            if any(x in lk for x in ("48", "premiere", "première", "action", "etape", "étape", "first")):
                log.info("  %r => len=%s", k, len(flat[k]))
        if args.xlsx and args.xlsx.is_file():
            _print_xlsx_headers(args.xlsx)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
