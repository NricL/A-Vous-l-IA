#!/usr/bin/env python3
"""
Generate static Avoulia parcours pages from the Excel case database.

Key guarantees:
- deterministic prompt pack per case (no LLM call at generation time)
- one route-ready HTML page per use_case_id
- QA report for missing data / fallbacks
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from jinja2 import Environment
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
DEFAULT_XLSX = BACKEND_DIR / "documents" / "base_cas_usage_v453.xlsx"
DEFAULT_MAPPING = BACKEND_DIR / "app" / "static" / "parcours" / "mapping_uc_hash.csv"
DEFAULT_OUTPUT_DIR = BACKEND_DIR / "app" / "static" / "parcours"

CASE_HEADERS = {
    "use_case_id",
    "cas_utilisation",
    "domaine_label",
    "intention",
    "description_cas_utilisation",
    "questions_qualification",
    "prerequis_donnees",
    "guardrails",
    "premiere_action_48h",
    "sensibilite_donnees",
    "secteur",
    "effort",
}

PAGE_TEMPLATE = """<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="robots" content="noindex, nofollow, noarchive">
  <title>Passer a l'action - {{ case.cas_utilisation }}</title>
  <style>
  :root{
    --ink:#16213A; --muted:#5A6478; --paper:#F5F7F4; --card:#FFFFFF;
    --brand:#2B59C3; --ok:#1E7A52; --line:#DFE3DB; --warn:#FBF6E5;
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--paper);color:var(--ink);font:16px/1.55 "Segoe UI",system-ui,sans-serif}
  .page{max-width:900px;margin:0 auto;padding:32px 18px 84px}
  h1{margin:10px 0;font-size:clamp(28px,4vw,42px);line-height:1.1}
  h2{margin:0 0 10px;font-size:22px}
  p{margin:10px 0}
  .meta{display:flex;flex-wrap:wrap;gap:8px;margin:14px 0 20px}
  .pill{background:#fff;border:1px solid var(--line);border-radius:999px;padding:5px 11px;font-size:13px;font-weight:600}
  .pill.warn{background:var(--warn)}
  .block{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:16px;margin:14px 0}
  .block ul{margin:8px 0 0 20px}
  .block li{margin:6px 0}
  .quickwin{border-left:4px solid var(--brand)}
  .progress{position:sticky;top:0;background:var(--paper);padding:12px 0;z-index:2}
  .progress-head{display:flex;justify-content:space-between;font-size:13px;font-weight:700}
  .bar{height:8px;border-radius:999px;background:#d8ddd5;margin-top:7px;overflow:hidden}
  .bar>div{height:100%;width:0;background:var(--ok);transition:width .25s ease}
  .step{position:relative;padding-left:52px;margin:14px 0}
  .dot{position:absolute;left:0;top:0;width:36px;height:36px;border-radius:999px;border:2px solid var(--line);display:flex;align-items:center;justify-content:center;background:#fff;font-weight:700}
  .step.done .dot{background:var(--ok);border-color:var(--ok);color:#fff}
  .step.done .dot::after{content:"✓"}
  .step.done .dot span{display:none}
  .step .block{margin:0}
  .sub{color:var(--muted);font-size:14px}
  .done{display:flex;gap:10px;align-items:flex-start;border-top:1px dashed var(--line);margin-top:12px;padding-top:12px;font-weight:700}
  .done input{margin-top:3px;accent-color:var(--ok)}
  .diag-row{display:grid;grid-template-columns:1fr auto;gap:8px;border-bottom:1px dashed var(--line);padding:10px 0}
  .diag-row:last-child{border-bottom:0}
  .diag-choices{display:flex;gap:12px;font-size:14px}
  .diag-status{margin-top:12px;padding:10px 12px;border-radius:8px;background:#eef4ff;font-weight:700}
  .ticket{background:#111a2e;color:#e8ecf5;border-radius:10px;overflow:hidden;margin-top:10px}
  .ticket-head{display:flex;justify-content:space-between;align-items:center;padding:8px 12px;border-bottom:1px dashed rgba(255,255,255,.25);font-size:12px;font-weight:700;text-transform:uppercase}
  .ticket pre{margin:0;padding:12px;white-space:pre-wrap;word-break:break-word;font:13px/1.6 Consolas,monospace}
  .copy{background:var(--brand);color:#fff;border:0;border-radius:8px;padding:6px 10px;font-size:12px;font-weight:700;cursor:pointer}
  footer{margin-top:42px;color:var(--muted);font-size:13px}
  </style>
</head>
<body>
<main class="page">
  <p class="sub">Parcours de mise en oeuvre · Avoulia</p>
  <h1>{{ case.cas_utilisation }}</h1>
  <p>{{ case.description_cas_utilisation }}</p>
  <div class="meta">
    <span class="pill">{{ case.domaine_label }}</span>
    <span class="pill">{{ case.intention }}</span>
    <span class="pill">Effort : {{ case.effort }}</span>
    <span class="pill warn">Donnees : {{ case.sensibilite_donnees }}</span>
    <span class="pill">Secteur : {{ case.secteur }}</span>
  </div>

  <section class="block">
    <h2>Ce parcours vous permet de</h2>
    <ul>
      <li>Un cadrage operationnel sur votre cas d'usage.</li>
      <li>Un livrable testable en 48h.</li>
      <li>Un protocole de test et de mise en routine.</li>
    </ul>
  </section>

  <section class="block">
    <h2>Comment utiliser les prompts</h2>
    <ol>
      <li>Ouvrez l'assistant IA de votre choix (celui que vous utilisez deja).</li>
      <li>Copiez/collez le prompt de l'etape en cours, puis ajoutez vos informations.</li>
      <li>Le livrable genere a chaque etape devient l'entree de l'etape suivante.</li>
      <li>Conservez vos 5 documents: ils serviront de trace de travail partageable.</li>
    </ol>
  </section>

  <div class="progress" aria-live="polite">
    <div class="progress-head"><span>Avancement</span><span id="count">0 etape sur 6</span></div>
    <div class="bar"><div id="fill"></div></div>
  </div>

  <section class="step" data-step="1">
    <div class="dot"><span>1</span></div>
    <div class="block">
      <h2>Diagnostic rapide (2 min)</h2>
      <p class="sub">Cette etape vous aide a choisir un demarrage immediat ou progressif.</p>
      {% for q in case.questions_qualification_items %}
      <div class="diag-row">
        <div>{{ q }}</div>
        <div class="diag-choices">
          <label><input id="d{{ loop.index }}-oui" type="radio" name="d{{ loop.index }}" value="oui"> Oui</label>
          <label><input id="d{{ loop.index }}-partiel" type="radio" name="d{{ loop.index }}" value="partiel"> Partiel</label>
          <label><input id="d{{ loop.index }}-non" type="radio" name="d{{ loop.index }}" value="non"> Non</label>
        </div>
      </div>
      {% endfor %}
      <div id="diag-status" class="diag-status">Decision : repondez aux questions pour obtenir une recommandation.</div>
      <div class="done">
        <input id="fin-1" type="checkbox" data-fin="1">
        <label for="fin-1">Etape terminee : diagnostic valide</label>
      </div>
    </div>
  </section>

  <section class="block quickwin">
    <h2>Demarrage guide</h2>
    <p>Utilisez ce prompt pour generer rapidement votre Document 1.</p>
    <div class="ticket">
      <div class="ticket-head"><span>Prompt etape 1 (Document 1)</span><button class="copy" data-copy="#quickwin-prompt">Copier</button></div>
      <pre id="quickwin-prompt">{{ prompts.quickwin_prompt }}</pre>
    </div>
  </section>

  <section class="step" data-step="2">
    <div class="dot"><span>2</span></div>
    <div class="block">
      <h2>Rassembler les informations utiles (30 min)</h2>
      <p><strong>Objectif :</strong> preparer les informations indispensables avant production.</p>
      <p class="sub"><strong>Entree etape 2 :</strong> reponses au diagnostic et contexte metier.</p>
      <ul>{% for item in case.prerequis_items %}<li>{{ item }}</li>{% endfor %}</ul>
      <div class="ticket">
        <div class="ticket-head"><span>Prompt etape 2 (Document 2)</span><button class="copy" data-copy="#step2-prompt">Copier</button></div>
        <pre id="step2-prompt">{{ prompts.step2_prepare_prompt }}</pre>
      </div>
      <p><strong>Livrable :</strong> une note de contexte claire pour lancer la production.</p>
      <p class="sub"><strong>Sortie attendue :</strong> Document 2 - Dossier de preparation (a reutiliser tel quel en etape 4).</p>
      <p><strong>C'est fait quand :</strong> un tiers comprend votre besoin sans explication orale.</p>
      <div class="done"><input id="fin-2" type="checkbox" data-fin="2"><label for="fin-2">Etape terminee</label></div>
    </div>
  </section>

  <section class="step" data-step="3">
    <div class="dot"><span>3</span></div>
    <div class="block">
      <h2>Poser le cadre de prudence (10 min)</h2>
      <p class="sub">Definissez les regles de prudence avant de produire.</p>
      <ul>{% for item in case.guardrails_items %}<li>{{ item }}</li>{% endfor %}</ul>
      <p class="sub">Niveau de sensibilite : {{ case.sensibilite_donnees }}.</p>
      <div class="done"><input id="fin-3" type="checkbox" data-fin="3"><label for="fin-3">Etape terminee</label></div>
    </div>
  </section>

  <section class="step" data-step="4">
    <div class="dot"><span>4</span></div>
    <div class="block">
      <h2>Produire le premier livrable (1 h)</h2>
      <p class="sub"><strong>Entree etape 4 :</strong> Document 2 - Dossier de preparation valide.</p>
      <p><strong>Action de reference (48h) :</strong> {{ case.premiere_action_48h_compact }}</p>
      <div class="ticket">
        <div class="ticket-head"><span>Prompt etape 4 (Document 3)</span><button class="copy" data-copy="#step4-prompt">Copier</button></div>
        <pre id="step4-prompt">{{ prompts.step4_execute_prompt }}</pre>
      </div>
      <p><strong>C'est fait quand :</strong> un premier livrable complet est relu en interne.</p>
      <p class="sub"><strong>Sortie attendue :</strong> Document 3 - Livrable operationnel (base du test terrain etape 5).</p>
      <div class="done"><input id="fin-4" type="checkbox" data-fin="4"><label for="fin-4">Etape terminee</label></div>
    </div>
  </section>

  <section class="step" data-step="5">
    <div class="dot"><span>5</span></div>
    <div class="block">
      <h2>Tester sur un petit perimetre (semaine en cours)</h2>
      <p class="sub"><strong>Entree etape 5 :</strong> Document 3 - Livrable operationnel.</p>
      <div class="ticket">
        <div class="ticket-head"><span>Prompt etape 5 (Document 4)</span><button class="copy" data-copy="#step5-prompt">Copier</button></div>
        <pre id="step5-prompt">{{ prompts.step5_test_prompt }}</pre>
      </div>
      <p><strong>C'est fait quand :</strong> vous savez quoi garder, ajuster ou abandonner.</p>
      <p class="sub"><strong>Sortie attendue :</strong> Document 4 - Resultat de test (resultats + decision).</p>
      <div class="done"><input id="fin-5" type="checkbox" data-fin="5"><label for="fin-5">Etape terminee</label></div>
    </div>
  </section>

  <section class="step" data-step="6">
    <div class="dot"><span>6</span></div>
    <div class="block">
      <h2>Installer la routine d'equipe (30 min)</h2>
      <p class="sub"><strong>Entree etape 6 :</strong> Document 4 - Resultat de test.</p>
      <div class="ticket">
        <div class="ticket-head"><span>Prompt etape 6 (Document 5)</span><button class="copy" data-copy="#step6-prompt">Copier</button></div>
        <pre id="step6-prompt">{{ prompts.step6_operate_prompt }}</pre>
      </div>
      <p><strong>C'est fait quand :</strong> l'usage tourne sans dependre d'une seule personne.</p>
      <p class="sub"><strong>Sortie attendue :</strong> Document 5 - Mode operatoire equipe partageable.</p>
      <div class="done"><input id="fin-6" type="checkbox" data-fin="6"><label for="fin-6">Etape terminee</label></div>
    </div>
  </section>

  <footer>
    Cas {{ case.use_case_id }} · genere automatiquement a partir de la base Avoulia.
  </footer>
</main>

<script>
const pageKey = "avoulia-parcours-{{ case.case_hash }}";
const finChecks = Array.from(document.querySelectorAll("[data-fin]"));
const fill = document.getElementById("fill");
const count = document.getElementById("count");
const diagStatus = document.getElementById("diag-status");
const diagCount = {{ case.questions_qualification_items|length }};

function loadState() {
  const raw = localStorage.getItem(pageKey);
  if (!raw) return {};
  try { return JSON.parse(raw); } catch { return {}; }
}
function saveState(state) {
  localStorage.setItem(pageKey, JSON.stringify(state));
}
function refreshProgress() {
  const done = finChecks.filter(i => i.checked).length;
  const pct = Math.round((done / 6) * 100);
  fill.style.width = pct + "%";
  count.textContent = done + " etape" + (done > 1 ? "s" : "") + " sur 6";
  document.querySelectorAll(".step").forEach(step => {
    const n = step.getAttribute("data-step");
    const fin = document.querySelector("#fin-" + n);
    step.classList.toggle("done", Boolean(fin && fin.checked));
  });
}
function refreshDiag() {
  const answers = [];
  for (let i = 1; i <= diagCount; i++) {
    const checked = document.querySelector('input[name="d' + i + '"]:checked');
    answers.push(checked ? checked.value : null);
  }
  if (answers.includes(null)) {
    diagStatus.textContent = "Decision : repondez aux " + diagCount + " questions pour obtenir une recommandation.";
    return;
  }
  const score = answers.reduce((s, v) => s + (v === "oui" ? 2 : v === "partiel" ? 1 : 0), 0);
  diagStatus.textContent = score >= (diagCount + 1)
    ? "Decision : GO maintenant (vous pouvez lancer ce parcours aujourd'hui)."
    : "Decision : GO progressif (commencez sur petit perimetre puis elargissez).";
}

const state = loadState();
finChecks.forEach(cb => {
  cb.checked = Boolean(state[cb.id]);
  cb.addEventListener("change", () => {
    state[cb.id] = cb.checked;
    saveState(state);
    refreshProgress();
  });
});
document.querySelectorAll('input[type="radio"][name^="d"]').forEach(r => {
  if (state[r.id]) r.checked = true;
  r.addEventListener("change", () => {
    document.querySelectorAll('input[name="' + r.name + '"]').forEach(x => { state[x.id] = x.checked; });
    saveState(state);
    refreshDiag();
  });
});
document.querySelectorAll("[data-copy]").forEach(btn => {
  btn.addEventListener("click", async () => {
    const target = document.querySelector(btn.getAttribute("data-copy"));
    if (!target) return;
    await navigator.clipboard.writeText(target.textContent.trim());
    btn.textContent = "Copie !";
    setTimeout(() => { btn.textContent = "Copier"; }, 1200);
  });
});
refreshProgress();
refreshDiag();
</script>
</body>
</html>
"""


@dataclass
class GenerationStats:
    generated: int = 0
    fallback_questions: int = 0
    fallback_prerequis: int = 0
    fallback_guardrails: int = 0
    fallback_step4: int = 0
    qa_issues: int = 0


def split_items(raw: str, fallback: list[str]) -> tuple[list[str], bool]:
    text = (raw or "").strip()
    if not text:
        return fallback, True
    parts = [p.strip(" -\t\r\n") for p in re.split(r"\s*\|\s*|\n+|;\s*", text) if p and p.strip()]
    clean = [p for p in parts if p]
    if not clean:
        return fallback, True
    return clean, False


def clean_text(value: Any) -> str:
    return str(value or "").replace("\r", " ").replace("\n", " ").strip()


def parse_cases_from_xlsx(xlsx_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = [clean_text(c) for c in next(rows)]
    missing = CASE_HEADERS - set(headers)
    if missing:
        wb.close()
        raise ValueError(f"Missing Excel headers: {sorted(missing)}")

    idx = {h: i for i, h in enumerate(headers)}
    out: list[dict[str, str]] = []
    for row in rows:
        case_id = clean_text(row[idx["use_case_id"]])
        if not case_id:
            continue
        case: dict[str, str] = {}
        for h in CASE_HEADERS:
            case[h] = clean_text(row[idx[h]])
        out.append(case)
    wb.close()
    return out


def load_mapping(mapping_path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    with mapping_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            case_id = clean_text(row.get("case_id", "")).upper()
            case_hash = clean_text(row.get("case_hash", "")).lower()
            if case_id and case_hash:
                out[case_id] = case_hash
    return out


def sha_fallback_hash(case_id: str, salt: str) -> str:
    return hashlib.sha256(f"{case_id}|{salt}".encode("utf-8")).hexdigest()[:16]


def compact_step4_action(raw_action: str) -> str:
    """
    Réduit la première action 48h en version opérable:
    - garde l'intention cœur
    - supprime les longs blocs d'astuces/exemples
    """
    action = (raw_action or "").strip()
    if not action:
        return ""
    action = re.split(r"\bAstuce\b\s*:", action, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    action = re.split(r"\bVérifier\b\s*:", action, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    action = re.split(r"\bLivrable\b\s*:", action, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    action = re.sub(
        r"Ouvrir un assistant IA conversationnel gratuit en lui décrivant votre besoin\.?",
        "",
        action,
        flags=re.IGNORECASE,
    ).strip()
    action = re.sub(r"\s+", " ", action).strip(" .")
    return action + "."


def build_prompt_pack(case: dict[str, Any]) -> dict[str, str]:
    question_items: list[str] = case["questions_qualification_items"]
    prereq_items: list[str] = case["prerequis_items"]
    guardrails_items: list[str] = case["guardrails_items"]

    p_lines = "\n".join(f"- {p}" for p in prereq_items)
    g_lines = "\n".join(f"- {g}" for g in guardrails_items)

    step4_action = compact_step4_action(case["premiere_action_48h"])
    if not step4_action:
        step4_action = (
            "Construis un premier livrable concret base sur les informations disponibles, "
            "puis propose les donnees minimales manquantes."
        )

    context_header = (
        "Je travaille dans une entreprise.\n"
        f"Ma fonction se situe dans le domaine metier: {case['domaine_label']}.\n"
        f"Mon objectif prioritaire est: {case['intention']}.\n"
        f"Contexte secteur: {case['secteur']}.\n"
        f"Cas d'usage cible: {case['cas_utilisation']}.\n"
    )

    quickwin_prompt = (
        context_header
        + "\nMISSION: produire un premier livrable actionnable en moins de 48h.\n"
        "ENTREE: contexte metier ci-dessus + reponses de diagnostic D1/D2/D3.\n"
        "SORTIE: Document 1 - Brief de depart en 1 page (probleme, proposition, plan 48h, risques).\n"
        "0) Verifie d'abord si D1/D2/D3 sont deja fournis dans mon message.\n"
        "0bis) Si D1/D2/D3 manquent, demande-moi ces 3 reponses puis attends ma reponse avant de continuer.\n"
        "1) Reprends les reponses du diagnostic et transforme-les en hypothese de travail.\n"
        "2) Si une information manque, liste-la explicitement dans une section 'Hypotheses' (sans inventer).\n"
        "3) Produis le Document 1 - Brief de depart directement utilisable.\n"
        "4) Termine par une checklist de verification avant diffusion.\n"
        "Format de sortie attendu: 4 blocs avec titres exacts: 'Probleme', 'Proposition', 'Plan 48h', 'Risques', puis 'Checklist'."
    )

    step2_prepare_prompt = (
        context_header
        + "\nMISSION: preparer les elements d'entree de production.\n"
        "ENTREE: contexte metier + Document 1 - Brief de depart + liste des prerequis.\n"
        "SORTIE: Document 2 - Dossier de preparation (tableau Element / Etat / Action immediate / Proprietaire / Echeance).\n"
        "Si le Document 1 - Brief de depart est absent, demande-le d'abord a l'utilisateur puis attends sa reponse.\n"
        "Agis comme un facilitateur operationnel.\n"
        "Informations a rassembler:\n"
        + p_lines
        + "\nSi un element manque, pose les questions minimales puis propose une version provisoire exploitable.\n"
        "Format de sortie attendu: tableau markdown strict avec 5 colonnes: Element | Etat | Action immediate | Proprietaire | Echeance."
    )

    step4_execute_prompt = (
        context_header
        + "\nMISSION: executer la premiere action 48h du cas, de facon concrete.\n"
        "ENTREE: Document 2 - Dossier de preparation valide + action 48h.\n"
        "Si le Document 2 - Dossier de preparation est absent, demande-le d'abord a l'utilisateur puis attends sa reponse.\n"
        "ACTION_48H:\n"
        f"{step4_action}\n"
        "SORTIE: Document 3 - Livrable operationnel en 4 sections (contenu principal, checklist de verification, risques, plan J+2).\n"
        "Ne produis pas de plan theorique: fournis un livrable prêt a relire.\n"
        "\nContraintes de prudence a respecter:\n"
        + g_lines
        + "\nFormat de sortie attendu: 4 sections numerotees avec titres exacts: 1) Contenu principal 2) Checklist de verification 3) Risques 4) Plan J+2."
    )

    step5_test_prompt = (
        context_header
        + "\nMISSION: tester le Document 3 - Livrable operationnel sur un petit perimetre reel.\n"
        "ENTREE: Document 3 - Livrable operationnel.\n"
        "Si le Document 3 - Livrable operationnel est absent, demande-le d'abord a l'utilisateur puis attends sa reponse.\n"
        "SORTIE: Document 4 - Resultat de test (protocole + mesures + decision).\n"
        "Propose:\n"
        "- un protocole de test sur 7 jours,\n"
        "- les metriques de suivi (adoption, qualite, temps, impact),\n"
        "- les seuils 'garder / ajuster / arreter'.\n"
        "Format de sortie attendu: tableau markdown jour par jour (Jour | Action | Mesure | Observation) + bloc final 'Decision' (Garder/Ajuster/Arreter)."
    )

    step6_operate_prompt = (
        context_header
        + "\nMISSION: transformer le Document 4 - Resultat de test en routine equipe.\n"
        "ENTREE: Document 4 - Resultat de test + enseignements terrain.\n"
        "Si le Document 4 - Resultat de test est absent, demande-le d'abord a l'utilisateur puis attends sa reponse.\n"
        "SORTIE: Document 5 - Mode operatoire equipe (SOP 1 page + roles + rituel mensuel).\n"
        "Aide-moi a installer ce fonctionnement durablement dans l'equipe.\n"
        "Produis:\n"
        "- un mode operatoire en 1 page,\n"
        "- roles et responsabilites,\n"
        "- routine de revue mensuelle,\n"
        "- plan de passation a une deuxieme personne.\n"
        "Format de sortie attendu: 4 sections avec titres exacts: 'Mode operatoire (1 page)', 'Roles et responsabilites', 'Routine mensuelle', 'Plan de passation'."
    )

    return {
        "quickwin_prompt": quickwin_prompt.strip(),
        "step2_prepare_prompt": step2_prepare_prompt.strip(),
        "step4_execute_prompt": step4_execute_prompt.strip(),
        "step5_test_prompt": step5_test_prompt.strip(),
        "step6_operate_prompt": step6_operate_prompt.strip(),
    }


def qa_case(case: dict[str, Any], prompts: dict[str, str]) -> list[str]:
    issues: list[str] = []
    required_prompt_keys = (
        "quickwin_prompt",
        "step2_prepare_prompt",
        "step4_execute_prompt",
        "step5_test_prompt",
        "step6_operate_prompt",
    )
    for key in required_prompt_keys:
        value = (prompts.get(key) or "").strip()
        if not value:
            issues.append(f"missing_prompt:{key}")
        if "{{" in value or "}}" in value or "[TODO]" in value:
            issues.append(f"unresolved_placeholder:{key}")

    if len(case["questions_qualification_items"]) < 3:
        issues.append("insufficient_questions")
    if len(case["prerequis_items"]) < 2:
        issues.append("weak_prerequis")
    return issues


def prepare_case(case: dict[str, str], mapping: dict[str, str], salt: str, stats: GenerationStats) -> dict[str, Any]:
    case_id = case["use_case_id"].upper()
    case_hash = mapping.get(case_id) or sha_fallback_hash(case_id, salt)

    question_items, q_fallback = split_items(
        case["questions_qualification"],
        [
            "Le besoin est-il prioritaire cette semaine ?",
            "Avez-vous un sponsor interne pour avancer ?",
            "Disposez-vous d'un minimum de donnees ou d'exemples ?",
        ],
    )
    prereq_items, p_fallback = split_items(
        case["prerequis_donnees"],
        [
            "Description du contexte de depart",
            "Exemple concret recent",
            "Critere de succes attendu",
        ],
    )
    guardrails_items, g_fallback = split_items(
        case["guardrails"],
        [
            "Faire relire tout livrable avant diffusion.",
            "Ne pas exposer de donnees personnelles identifiables.",
            "Verifier les faits et chiffres avant publication.",
        ],
    )

    if q_fallback:
        stats.fallback_questions += 1
    if p_fallback:
        stats.fallback_prerequis += 1
    if g_fallback:
        stats.fallback_guardrails += 1
    compact_action = compact_step4_action(case["premiere_action_48h"])
    if not compact_action:
        stats.fallback_step4 += 1

    return {
        **case,
        "case_hash": case_hash,
        "questions_qualification_items": question_items[:3],
        "prerequis_items": prereq_items,
        "guardrails_items": guardrails_items,
        "premiere_action_48h_compact": compact_action
        or "Produire un premier livrable concret en 48h a partir des informations disponibles.",
    }


def render_page(env: Environment, case: dict[str, Any], prompts: dict[str, str]) -> str:
    tpl = env.from_string(PAGE_TEMPLATE)
    return tpl.render(case=case, prompts=prompts)


def generate_pages(
    xlsx_path: Path,
    mapping_path: Path,
    output_dir: Path,
    limit: int | None,
    preview_subdir: str,
    salt: str,
    qa_report_path: Path,
) -> GenerationStats:
    stats = GenerationStats()
    env = Environment(autoescape=True, trim_blocks=True, lstrip_blocks=True)
    mapping = load_mapping(mapping_path)
    cases = parse_cases_from_xlsx(xlsx_path)

    if limit is not None:
        cases = cases[:limit]

    target_dir = output_dir / preview_subdir if preview_subdir else output_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    qa_rows: list[dict[str, str]] = []

    for case in cases:
        prepared = prepare_case(case, mapping, salt, stats)
        prompts = build_prompt_pack(prepared)
        issues = qa_case(prepared, prompts)
        if issues:
            stats.qa_issues += 1

        page_html = render_page(env, prepared, prompts)
        out_file = target_dir / f"action-{prepared['case_hash']}.html"
        out_file.write_text(page_html, encoding="utf-8")
        stats.generated += 1

        qa_rows.append(
            {
                "use_case_id": prepared["use_case_id"],
                "case_hash": prepared["case_hash"],
                "issues": "|".join(issues),
                "fallback_questions": str(int(len(prepared["questions_qualification_items"]) < 3)),
                "fallback_prerequis": str(int(not prepared["prerequis_donnees"])),
                "fallback_guardrails": str(int(not prepared["guardrails"])),
                "fallback_step4": str(int(not prepared["premiere_action_48h"])),
            }
        )

    qa_report_path.parent.mkdir(parents=True, exist_ok=True)
    with qa_report_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "use_case_id",
                "case_hash",
                "issues",
                "fallback_questions",
                "fallback_prerequis",
                "fallback_guardrails",
                "fallback_step4",
            ],
        )
        writer.writeheader()
        writer.writerows(qa_rows)

    return stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate static Avoulia parcours pages.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Source Excel file path")
    parser.add_argument("--mapping", type=Path, default=DEFAULT_MAPPING, help="CSV mapping path (case_id -> case_hash)")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Output directory")
    parser.add_argument("--limit", type=int, default=None, help="Generate only first N cases")
    parser.add_argument(
        "--preview-subdir",
        type=str,
        default="",
        help="Optional subdirectory under output-dir (example: preview-v2)",
    )
    parser.add_argument(
        "--qa-report",
        type=Path,
        default=BACKEND_DIR / "app" / "static" / "parcours" / "qa_prompt_pack_report.csv",
        help="QA report CSV path",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    salt = (os.getenv("AVOULIA_SALT") or "dev-salt-12345").strip()

    if not args.xlsx.exists():
        raise FileNotFoundError(f"Excel file not found: {args.xlsx}")
    if not args.mapping.exists():
        raise FileNotFoundError(f"Mapping CSV not found: {args.mapping}")

    stats = generate_pages(
        xlsx_path=args.xlsx,
        mapping_path=args.mapping,
        output_dir=args.output_dir,
        limit=args.limit,
        preview_subdir=args.preview_subdir.strip("/\\"),
        salt=salt,
        qa_report_path=args.qa_report,
    )

    summary = {
        "generated": stats.generated,
        "fallback_questions": stats.fallback_questions,
        "fallback_prerequis": stats.fallback_prerequis,
        "fallback_guardrails": stats.fallback_guardrails,
        "fallback_step4": stats.fallback_step4,
        "qa_issues": stats.qa_issues,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
